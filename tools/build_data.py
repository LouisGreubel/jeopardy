#!/usr/bin/env python3
"""Convert clues.xlsx into validated, sharded JSON for the Jeopardy web app.

The converter intentionally keeps complete five-clue category instances together,
filters high-confidence media-dependent clues, preserves Excel's displayed values,
and produces human-readable review reports.
"""

from __future__ import annotations

import argparse
import calendar
import csv
import hashlib
import json
import re
import shutil
import sys
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime, time, timezone
from pathlib import Path
from typing import Any, Iterable, Sequence

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - friendly command-line failure
    raise SystemExit(
        "openpyxl is required. Run: python -m pip install -r requirements.txt"
    ) from exc

CONVERTER_VERSION = "1.0.1"
SCHEMA_VERSION = 2
EXPECTED_HEADERS = ["round", "clue_value", "category", "clue", "response"]
ROUND_VALUES: dict[int, list[int]] = {
    1: [200, 400, 600, 800, 1000],
    2: [400, 800, 1200, 1600, 2000],
}
FINAL_VALUE = 3000
DEFAULT_SHARD_SIZE = 400

# These patterns are deliberately narrow. They exclude clues only when the clue
# text strongly indicates that missing media is part of the prompt. Broader
# possible-media wording is still reported for review, but remains playable.
MEDIA_PATTERNS: list[tuple[str, str, re.Pattern[str]]] = [
    (
        "monitor_visual",
        "The clue explicitly refers to material on a monitor.",
        re.compile(r"\bon (?:the|your) monitor\b", re.IGNORECASE),
    ),
    (
        "starts_with_here_is_or_are",
        "The clue begins by presenting an object or people that are not in the text.",
        re.compile(
            r"^\s*(?:\([^)]*\)\s*)*here (?:is|are|was|were)\b",
            re.IGNORECASE,
        ),
    ),
    (
        "shown_here",
        "The clue explicitly says something is shown, seen, or pictured here.",
        re.compile(
            r"\b(?:shown|seen|pictured|displayed|depicted|illustrated) "
            r"(?:here|above|below)\b",
            re.IGNORECASE,
        ),
    ),
    (
        "photo_or_image_reference",
        "The clue requires information contained in a photo or image.",
        re.compile(
            r"\b(?:in|from) (?:this|the) (?:photo|photograph|image)\b|"
            r"\b(?:this|the) (?:photo|photograph|image) "
            r"(?:shows|depicts|pictures|features)\b",
            re.IGNORECASE,
        ),
    ),
    (
        "map_or_diagram_prompt",
        "The clue directs the player to a missing map, diagram, or picture.",
        re.compile(
            r"\b(?:see|look at) (?:this|the) "
            r"(?:map|photo|photograph|picture|image|diagram|illustration)\b|"
            r"\bindicated on (?:this|the) map\b|"
            r"\bthe map here\b",
            re.IGNORECASE,
        ),
    ),
    (
        "audio_prompt",
        "The clue asks the player to listen to missing audio.",
        re.compile(
            r"(?:^|[.!?]\s+)(?:take a listen|listen\s*:)|"
            r"\baudio (?:clue|daily double)\b|"
            r"\bheard here\b|\byou just heard\b",
            re.IGNORECASE,
        ),
    ),
    (
        "bracketed_audio",
        "The transcript contains a bracketed audio or music cue.",
        re.compile(
            r"\[[^\]]*\b(?:plays?|music|sound|audio|singing|sings?|"
            r"revving|heard|voice|speaks?)\b[^\]]*\]",
            re.IGNORECASE,
        ),
    ),
    (
        "video_prompt",
        "The clue directs the player to watch missing video or footage.",
        re.compile(
            r"\b(?:watch|view) (?:this|the following) "
            r"(?:clip|video|footage)\b|"
            r"\b(?:video|film) clip (?:shows|of|from)\b",
            re.IGNORECASE,
        ),
    ),
    (
        "explicit_visual_object",
        "The clue identifies a person or object as being pictured here.",
        re.compile(
            r"\b(?:the|this) (?:man|woman|person|people|animal|building|"
            r"painting|sculpture|object|flag|logo|landmark|plant|vehicle) "
            r"(?:shown|seen|pictured) here\b",
            re.IGNORECASE,
        ),
    ),
]

SOFT_MEDIA_PATTERN = re.compile(
    r"\b(?:here|photo|photograph|image|picture|map|diagram|illustration|"
    r"animation|footage|listen|watch)\b|\[[^\]]+\]",
    re.IGNORECASE,
)


@dataclass(slots=True)
class CellOverride:
    column: str
    start_row: int
    end_row: int
    value: str
    reason: str

    def applies(self, column: str, row_number: int) -> bool:
        return self.column == column and self.start_row <= row_number <= self.end_row


@dataclass(slots=True)
class RowRecord:
    source_row: int
    round_key: int | str
    value: int
    category: str
    clue: str
    response: str
    category_key: str
    high_media_reasons: list[str] = field(default_factory=list)
    soft_media_review: bool = False
    field_errors: list[str] = field(default_factory=list)


@dataclass(slots=True)
class CategoryCandidate:
    round_number: int
    category: str
    records: list[RowRecord]

    @property
    def source_rows(self) -> list[int]:
        return [record.source_row for record in self.records]


class ConversionError(RuntimeError):
    """Raised when the workbook structure is incompatible with the converter."""


def parse_args() -> argparse.Namespace:
    script_path = Path(__file__).resolve()
    project_root = script_path.parent.parent
    parser = argparse.ArgumentParser(
        description="Build validated, sharded Jeopardy clue data from clues.xlsx."
    )
    parser.add_argument(
        "workbook",
        nargs="?",
        type=Path,
        default=project_root / "source" / "clues.xlsx",
        help="Path to clues.xlsx (default: source/clues.xlsx).",
    )
    parser.add_argument(
        "--project-root",
        type=Path,
        default=project_root,
        help="Project root that contains data/, docs/, reports/, and tools/.",
    )
    parser.add_argument(
        "--shard-size",
        type=int,
        default=DEFAULT_SHARD_SIZE,
        help=f"Records per JSON shard (default: {DEFAULT_SHARD_SIZE}).",
    )
    parser.add_argument(
        "--overrides",
        type=Path,
        default=script_path.parent / "data_overrides.json",
        help="JSON file containing explicit Excel-conversion repairs.",
    )
    parser.add_argument(
        "--include-high-confidence-media",
        action="store_true",
        help="Keep clues that explicitly depend on missing media. Not recommended.",
    )
    return parser.parse_args()


def normalize_text(value: str) -> str:
    """Normalize Unicode and whitespace without changing punctuation or wording."""
    text = unicodedata.normalize("NFC", value)
    text = text.replace("\u00a0", " ").replace("\r\n", "\n").replace("\r", "\n")
    # The browser collapses ordinary whitespace. Normalizing it here also makes
    # stable IDs independent of accidental Excel spacing.
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def normalized_key(value: str) -> str:
    return unicodedata.normalize("NFKC", value).casefold().strip()


def decimal_places_from_format(number_format: str, marker: str | None = None) -> int:
    fmt = number_format
    if marker and marker in fmt:
        fmt = fmt.split(marker, 1)[0]
    match = re.search(r"\.([0#]+)", fmt)
    return len(match.group(1)) if match else 0


def format_number(value: int | float, number_format: str) -> str:
    fmt = number_format or "General"
    lower_fmt = fmt.lower()

    if "%" in fmt:
        decimals = decimal_places_from_format(fmt, "%")
        return f"{float(value) * 100:.{decimals}f}%"

    is_currency = "$" in fmt
    decimals = decimal_places_from_format(fmt)
    use_grouping = "," in fmt

    if fmt == "General":
        number = float(value)
        if number.is_integer():
            return str(int(number))
        return format(number, ".15g")

    if decimals == 0 and float(value).is_integer():
        numeric_text = f"{int(value):,}" if use_grouping else str(int(value))
    else:
        numeric_text = (
            f"{float(value):,.{decimals}f}"
            if use_grouping
            else f"{float(value):.{decimals}f}"
        )

    return f"${numeric_text}" if is_currency else numeric_text


def format_temporal(value: datetime | date | time, number_format: str) -> str:
    fmt = (number_format or "General").lower().replace("\\", "")

    if isinstance(value, time):
        hour = value.hour
        minute = value.minute
        if "am/pm" in fmt:
            suffix = "AM" if hour < 12 else "PM"
            hour_12 = hour % 12 or 12
            return f"{hour_12}:{minute:02d} {suffix}"
        return f"{hour}:{minute:02d}"

    if isinstance(value, datetime):
        date_value = value.date()
    else:
        date_value = value

    month = date_value.month
    day = date_value.day
    year = date_value.year
    month_name = calendar.month_name[month]

    if fmt == "m, d":
        return f"{month}, {day}"
    if fmt == "mmmm d":
        return f"{month_name} {day}"
    if fmt == "mmmm d, yyyy":
        return f"{month_name} {day}, {year}"
    if fmt == "mmmm yyyy":
        return f"{month_name} {year}"
    if fmt == "m/d":
        return f"{month}/{day}"
    if fmt == "m-d":
        return f"{month}-{day}"
    if fmt == "m-d-yyyy":
        return f"{month}-{day}-{year}"

    return date_value.isoformat()


def load_overrides(path: Path) -> list[CellOverride]:
    if not path.exists():
        return []

    payload = json.loads(path.read_text(encoding="utf-8"))
    raw_overrides = payload.get("cellOverrides", [])
    overrides: list[CellOverride] = []

    for item in raw_overrides:
        column = str(item["column"]).upper()
        if column not in {"A", "B", "C", "D", "E"}:
            raise ConversionError(f"Unsupported override column: {column}")
        start_row = int(item["startRow"])
        end_row = int(item.get("endRow", start_row))
        if end_row < start_row:
            raise ConversionError(
                f"Override endRow {end_row} precedes startRow {start_row}."
            )
        overrides.append(
            CellOverride(
                column=column,
                start_row=start_row,
                end_row=end_row,
                value=normalize_text(str(item["value"])),
                reason=normalize_text(str(item.get("reason", "Manual repair"))),
            )
        )

    return overrides


def find_override(
    overrides: Sequence[CellOverride], column: str, row_number: int
) -> CellOverride | None:
    for override in overrides:
        if override.applies(column, row_number):
            return override
    return None


def format_cell(
    cell: Any,
    column: str,
    row_number: int,
    overrides: Sequence[CellOverride],
) -> tuple[str, CellOverride | None]:
    override = find_override(overrides, column, row_number)
    if override:
        return override.value, override

    value = cell.value
    if value is None:
        return "", None
    if isinstance(value, str):
        return normalize_text(value), None
    if isinstance(value, bool):
        return ("true" if value else "false"), None
    if isinstance(value, (datetime, date, time)):
        return normalize_text(format_temporal(value, cell.number_format)), None
    if isinstance(value, (int, float)):
        return normalize_text(format_number(value, cell.number_format)), None
    return normalize_text(str(value)), None


def parse_round(value: Any) -> int | str:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        int_value = int(value)
        if int_value in (1, 2) and float(value) == int_value:
            return int_value

    text = normalize_text(str(value)).casefold() if value is not None else ""
    if text in {"final jeopardy!", "final jeopardy", "final"}:
        return "final"
    raise ConversionError(f"Unrecognized round value: {value!r}")


def parse_clue_value(value: Any) -> int:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        number = int(value)
        if float(value) == number:
            return number

    text = normalize_text(str(value)) if value is not None else ""
    text = text.replace("$", "").replace(",", "")
    if re.fullmatch(r"-?\d+(?:\.0+)?", text):
        return int(float(text))
    raise ConversionError(f"Unrecognized clue value: {value!r}")


def media_reasons(clue: str) -> tuple[list[str], bool]:
    reasons = [name for name, _description, pattern in MEDIA_PATTERNS if pattern.search(clue)]
    return reasons, bool(SOFT_MEDIA_PATTERN.search(clue)) and not reasons


def stable_digest(payload: Any) -> str:
    canonical = json.dumps(
        payload,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("utf-8")
    return hashlib.sha256(canonical).hexdigest()


def build_category_object(candidate: CategoryCandidate) -> dict[str, Any]:
    content_payload = {
        "round": candidate.round_number,
        "category": candidate.category,
        "clues": [
            {
                "value": record.value,
                "clue": record.clue,
                "response": record.response,
            }
            for record in candidate.records
        ],
    }
    category_digest = stable_digest(content_payload)
    category_id = f"r{candidate.round_number}-{category_digest[:16]}"

    clues: list[dict[str, Any]] = []
    for record in candidate.records:
        clue_payload = {
            "categoryId": category_id,
            "value": record.value,
            "clue": record.clue,
            "response": record.response,
        }
        clue_digest = stable_digest(clue_payload)
        clues.append(
            {
                "id": f"r{candidate.round_number}c-{clue_digest[:16]}",
                "value": record.value,
                "clue": record.clue,
                "response": record.response,
            }
        )

    return {"id": category_id, "category": candidate.category, "clues": clues}


def build_final_object(record: RowRecord) -> dict[str, Any]:
    content_payload = {
        "round": "final",
        "value": record.value,
        "category": record.category,
        "clue": record.clue,
        "response": record.response,
    }
    digest = stable_digest(content_payload)
    return {
        "id": f"fj-{digest[:16]}",
        "category": record.category,
        "value": record.value,
        "clue": record.clue,
        "response": record.response,
    }


def inspect_workbook(
    workbook_path: Path,
    overrides: Sequence[CellOverride],
) -> tuple[
    dict[int | str, list[RowRecord]],
    list[dict[str, Any]],
    list[dict[str, Any]],
    dict[str, Any],
]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if len(workbook.worksheets) != 1:
        raise ConversionError(
            f"Expected one worksheet; found {len(workbook.worksheets)}."
        )

    worksheet = workbook.active
    header_cells = next(worksheet.iter_rows(min_row=1, max_row=1, max_col=5))
    headers = [normalize_text(str(cell.value or "")).casefold() for cell in header_cells]
    if headers != EXPECTED_HEADERS:
        raise ConversionError(
            "Unexpected headers. Expected "
            f"{EXPECTED_HEADERS!r}; found {headers!r}."
        )

    records_by_round: dict[int | str, list[RowRecord]] = defaultdict(list)
    converted_cells: list[dict[str, Any]] = []
    media_review_rows: list[dict[str, Any]] = []
    source_round_counts: Counter[str] = Counter()
    source_value_counts: dict[str, Counter[int]] = defaultdict(Counter)
    total_rows = 0

    for row_number, cells in enumerate(
        worksheet.iter_rows(min_row=2, max_col=5), start=2
    ):
        if all(cell.value is None for cell in cells):
            continue

        total_rows += 1
        round_key = parse_round(cells[0].value)
        value = parse_clue_value(cells[1].value)
        category, category_override = format_cell(cells[2], "C", row_number, overrides)
        clue, clue_override = format_cell(cells[3], "D", row_number, overrides)
        response, response_override = format_cell(cells[4], "E", row_number, overrides)

        for column, cell, display_text, override in (
            ("C", cells[2], category, category_override),
            ("D", cells[3], clue, clue_override),
            ("E", cells[4], response, response_override),
        ):
            if override or cell.data_type != "s" or cell.number_format != "General":
                converted_cells.append(
                    {
                        "source_row": row_number,
                        "column": column,
                        "data_type": cell.data_type,
                        "number_format": cell.number_format,
                        "original_value": repr(cell.value),
                        "display_text": display_text,
                        "override_applied": "yes" if override else "no",
                        "override_reason": override.reason if override else "",
                    }
                )

        field_errors: list[str] = []
        if not category:
            field_errors.append("blank_category")
        if not clue:
            field_errors.append("blank_clue")
        if not response:
            field_errors.append("blank_response")

        high_reasons, soft_review = media_reasons(clue)
        record = RowRecord(
            source_row=row_number,
            round_key=round_key,
            value=value,
            category=category,
            clue=clue,
            response=response,
            category_key=normalized_key(category),
            high_media_reasons=high_reasons,
            soft_media_review=soft_review,
            field_errors=field_errors,
        )
        records_by_round[round_key].append(record)

        round_label = str(round_key)
        source_round_counts[round_label] += 1
        source_value_counts[round_label][value] += 1

        if high_reasons or soft_review:
            media_review_rows.append(
                {
                    "source_row": row_number,
                    "round": round_label,
                    "value": value,
                    "category": category,
                    "clue": clue,
                    "response": response,
                    "confidence": "high" if high_reasons else "review_only",
                    "reason_codes": ";".join(high_reasons),
                }
            )

    workbook.close()

    source_stats = {
        "worksheet": worksheet.title,
        "rowCount": total_rows,
        "roundCounts": dict(sorted(source_round_counts.items())),
        "valueCounts": {
            key: {str(value): count for value, count in sorted(counter.items())}
            for key, counter in sorted(source_value_counts.items())
        },
    }
    return records_by_round, converted_cells, media_review_rows, source_stats


def reconstruct_categories(
    records: Sequence[RowRecord],
    round_number: int,
) -> tuple[list[CategoryCandidate], list[dict[str, Any]]]:
    expected_values = ROUND_VALUES[round_number]
    candidates: list[CategoryCandidate] = []
    excluded_rows: list[dict[str, Any]] = []
    index = 0

    while index < len(records):
        window = list(records[index : index + 5])
        is_complete = (
            len(window) == 5
            and [record.value for record in window] == expected_values
            and len({record.category_key for record in window}) == 1
            and all(not record.field_errors for record in window)
        )

        if is_complete:
            candidates.append(
                CategoryCandidate(
                    round_number=round_number,
                    category=window[0].category,
                    records=window,
                )
            )
            index += 5
            continue

        record = records[index]
        reasons: list[str] = []
        if record.field_errors:
            reasons.extend(record.field_errors)
        if record.value != expected_values[0]:
            reasons.append("not_part_of_complete_five_value_sequence")
        else:
            reasons.append("incomplete_or_malformed_category_sequence")

        excluded_rows.append(
            {
                "source_row": record.source_row,
                "round": round_number,
                "value": record.value,
                "category": record.category,
                "reason_codes": ";".join(sorted(set(reasons))),
                "clue": record.clue,
                "response": record.response,
            }
        )
        index += 1

    return candidates, excluded_rows


def filter_and_deduplicate_categories(
    candidates: Iterable[CategoryCandidate],
    include_high_confidence_media: bool,
) -> tuple[
    list[dict[str, Any]],
    list[dict[str, Any]],
    list[dict[str, Any]],
]:
    accepted: list[dict[str, Any]] = []
    excluded_categories: list[dict[str, Any]] = []
    duplicate_records: list[dict[str, Any]] = []
    seen_ids: dict[str, CategoryCandidate] = {}

    for candidate in candidates:
        media_rows = [
            record
            for record in candidate.records
            if record.high_media_reasons
        ]
        if media_rows and not include_high_confidence_media:
            reason_codes = sorted(
                {
                    reason
                    for record in media_rows
                    for reason in record.high_media_reasons
                }
            )
            excluded_categories.append(
                {
                    "round": candidate.round_number,
                    "start_row": min(candidate.source_rows),
                    "end_row": max(candidate.source_rows),
                    "category": candidate.category,
                    "reason_codes": ";".join(reason_codes),
                    "media_rows": ";".join(str(record.source_row) for record in media_rows),
                }
            )
            continue

        category_object = build_category_object(candidate)
        existing = seen_ids.get(category_object["id"])
        if existing is not None:
            duplicate_records.append(
                {
                    "kind": f"round_{candidate.round_number}_category",
                    "id": category_object["id"],
                    "category": candidate.category,
                    "first_source_rows": ";".join(
                        str(value) for value in existing.source_rows
                    ),
                    "duplicate_source_rows": ";".join(
                        str(value) for value in candidate.source_rows
                    ),
                }
            )
            continue

        seen_ids[category_object["id"]] = candidate
        accepted.append(category_object)

    accepted.sort(key=lambda item: item["id"])
    return accepted, excluded_categories, duplicate_records


def filter_and_deduplicate_final(
    records: Iterable[RowRecord],
    include_high_confidence_media: bool,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    accepted: list[dict[str, Any]] = []
    excluded: list[dict[str, Any]] = []
    duplicates: list[dict[str, Any]] = []
    seen_ids: dict[str, RowRecord] = {}

    for record in records:
        reasons = list(record.field_errors)
        if record.value != FINAL_VALUE:
            reasons.append("unexpected_final_value")
        if record.high_media_reasons and not include_high_confidence_media:
            reasons.extend(record.high_media_reasons)

        if reasons:
            excluded.append(
                {
                    "source_row": record.source_row,
                    "round": "final",
                    "value": record.value,
                    "category": record.category,
                    "reason_codes": ";".join(sorted(set(reasons))),
                    "clue": record.clue,
                    "response": record.response,
                }
            )
            continue

        final_object = build_final_object(record)
        existing = seen_ids.get(final_object["id"])
        if existing is not None:
            duplicates.append(
                {
                    "kind": "final_clue",
                    "id": final_object["id"],
                    "category": record.category,
                    "first_source_rows": str(existing.source_row),
                    "duplicate_source_rows": str(record.source_row),
                }
            )
            continue

        seen_ids[final_object["id"]] = record
        accepted.append(final_object)

    accepted.sort(key=lambda item: item["id"])
    return accepted, excluded, duplicates


def clean_output_directories(project_root: Path) -> None:
    for directory in (
        project_root / "data" / "round1",
        project_root / "data" / "round2",
        project_root / "data" / "final",
    ):
        if directory.exists():
            shutil.rmtree(directory)
        directory.mkdir(parents=True, exist_ok=True)

    (project_root / "docs").mkdir(parents=True, exist_ok=True)
    (project_root / "reports").mkdir(parents=True, exist_ok=True)
    (project_root / "data").mkdir(parents=True, exist_ok=True)


def write_json(path: Path, payload: Any, *, pretty: bool = False) -> tuple[int, str]:
    if pretty:
        text = json.dumps(payload, ensure_ascii=False, indent=2) + "\n"
    else:
        text = json.dumps(
            payload,
            ensure_ascii=False,
            separators=(",", ":"),
        ) + "\n"
    encoded = text.encode("utf-8")
    path.write_bytes(encoded)
    return len(encoded), hashlib.sha256(encoded).hexdigest()


def write_shards(
    project_root: Path,
    items: Sequence[dict[str, Any]],
    folder: str,
    prefix: str,
    kind: str,
    shard_size: int,
) -> dict[str, Any]:
    output_directory = project_root / "data" / folder
    shards: list[dict[str, Any]] = []
    total_bytes = 0

    for shard_index, start in enumerate(range(0, len(items), shard_size)):
        shard_items = list(items[start : start + shard_size])
        filename = f"{prefix}-{shard_index:03d}.json"
        path = output_directory / filename
        payload = {
            "schemaVersion": SCHEMA_VERSION,
            "kind": kind,
            "items": shard_items,
        }
        byte_count, file_hash = write_json(path, payload)
        total_bytes += byte_count
        shards.append(
            {
                "path": f"{folder}/{filename}",
                "count": len(shard_items),
                "bytes": byte_count,
                "sha256": file_hash,
            }
        )

    return {
        "recordCount": len(items),
        "shardSize": shard_size,
        "shardCount": len(shards),
        "bytes": total_bytes,
        "shards": shards,
    }


def write_csv(path: Path, rows: Sequence[dict[str, Any]], fieldnames: Sequence[str]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def build_data_version(
    round_one: Sequence[dict[str, Any]],
    round_two: Sequence[dict[str, Any]],
    final_clues: Sequence[dict[str, Any]],
    overrides: Sequence[CellOverride],
    include_high_confidence_media: bool,
) -> str:
    digest = hashlib.sha256()
    configuration = {
        "schemaVersion": SCHEMA_VERSION,
        "converterVersion": CONVERTER_VERSION,
        "mediaPatternNames": [name for name, _description, _pattern in MEDIA_PATTERNS],
        "includeHighConfidenceMedia": include_high_confidence_media,
        "overrides": [
            {
                "column": override.column,
                "startRow": override.start_row,
                "endRow": override.end_row,
                "value": override.value,
                "reason": override.reason,
            }
            for override in overrides
        ],
    }
    digest.update(
        json.dumps(
            configuration,
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
        ).encode("utf-8")
    )
    for label, items in (
        ("round1", round_one),
        ("round2", round_two),
        ("final", final_clues),
    ):
        digest.update(label.encode("ascii"))
        for item in items:
            digest.update(
                json.dumps(
                    item,
                    ensure_ascii=False,
                    sort_keys=True,
                    separators=(",", ":"),
                ).encode("utf-8")
            )
    return digest.hexdigest()[:20]


def source_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file:
        for block in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def format_bytes(value: int) -> str:
    units = ["B", "KB", "MB", "GB"]
    size = float(value)
    for unit in units:
        if size < 1024 or unit == units[-1]:
            return f"{size:.1f} {unit}" if unit != "B" else f"{int(size)} B"
        size /= 1024
    return f"{value} B"


def write_markdown_report(
    path: Path,
    workbook_path: Path,
    source_hash: str,
    source_stats: dict[str, Any],
    round_one_candidates: int,
    round_two_candidates: int,
    round_one: Sequence[dict[str, Any]],
    round_two: Sequence[dict[str, Any]],
    final_clues: Sequence[dict[str, Any]],
    excluded_rows: Sequence[dict[str, Any]],
    excluded_categories: Sequence[dict[str, Any]],
    excluded_final: Sequence[dict[str, Any]],
    duplicates: Sequence[dict[str, Any]],
    media_review_rows: Sequence[dict[str, Any]],
    converted_cells: Sequence[dict[str, Any]],
    manifest: dict[str, Any],
) -> None:
    high_media_rows = sum(
        1 for row in media_review_rows if row["confidence"] == "high"
    )
    soft_media_rows = sum(
        1 for row in media_review_rows if row["confidence"] == "review_only"
    )
    total_data_bytes = manifest["totals"]["dataBytes"]

    report = f"""# Phase 3 data-conversion report

Generated from `{workbook_path.name}` by converter version {CONVERTER_VERSION}.

## Source workbook

- Worksheet: `{source_stats['worksheet']}`
- Clue rows: {source_stats['rowCount']:,}
- Round 1 rows: {source_stats['roundCounts'].get('1', 0):,}
- Round 2 rows: {source_stats['roundCounts'].get('2', 0):,}
- Final Jeopardy rows: {source_stats['roundCounts'].get('final', 0):,}
- Source SHA-256: `{source_hash}`

## Reconstructed category instances

- Complete Round 1 instances before media filtering and deduplication: {round_one_candidates:,}
- Complete Round 2 instances before media filtering and deduplication: {round_two_candidates:,}
- Rows not belonging to a complete five-value sequence: {len(excluded_rows):,}

## Playable output

- Round 1 categories: {len(round_one):,} ({len(round_one) * 5:,} clues)
- Round 2 categories: {len(round_two):,} ({len(round_two) * 5:,} clues)
- Final Jeopardy clues: {len(final_clues):,}
- Total playable clues: {len(round_one) * 5 + len(round_two) * 5 + len(final_clues):,}
- JSON data size before web-server compression: {format_bytes(total_data_bytes)}
- Data version: `{manifest['dataVersion']}`

## Exclusions and review queues

- Complete category instances excluded for high-confidence missing-media cues: {len(excluded_categories):,}
- Final clues excluded for validation or high-confidence media cues: {len(excluded_final):,}
- Exact duplicate category instances or Final clues removed: {len(duplicates):,}
- High-confidence media rows found: {high_media_rows:,}
- Additional possible-media rows retained but listed for review: {soft_media_rows:,}
- Excel-formatted or manually repaired cells recorded: {len(converted_cells):,}

The converter excludes only narrow, high-confidence media patterns. Broader clues mentioning words such as “photo,” “map,” “watch,” or “here” remain playable unless the text explicitly depends on missing media. Review `reports/media_review.csv` to tighten or relax that policy later.

## Generated files

- `data/manifest.json`: dataset version, counts, and shard locations
- `data/round1/*.json`: complete Round 1 category instances
- `data/round2/*.json`: complete Round 2 category instances
- `data/final/*.json`: Final Jeopardy clues
- `reports/conversion_summary.json`: machine-readable audit summary
- `reports/excluded_rows.csv`: incomplete or malformed regular-round rows
- `reports/excluded_category_instances.csv`: complete categories removed by the media policy
- `reports/excluded_final_clues.csv`: removed Final Jeopardy rows
- `reports/media_review.csv`: high-confidence and review-only media matches
- `reports/converted_cells.csv`: Excel-formatted values and explicit repairs
- `reports/duplicate_records.csv`: exact duplicates removed from the pool

## Explicit Excel repairs

The checked-in `tools/data_overrides.json` repairs the known cases where Excel changed the intended text:

- The category in source rows 2–6 is restored from numeric `2` to `2.0`.
- The responses in source rows 32,648 and 112,976 are restored to `9-1-1`.
- The response in source row 33,731 is restored to `3-1-1`.

Other number, currency, percentage, date, and time cells are converted according to their Excel number formats and recorded in `reports/converted_cells.csv`.
"""
    path.write_text(report, encoding="utf-8")


def main() -> int:
    args = parse_args()
    workbook_path = args.workbook.resolve()
    project_root = args.project_root.resolve()

    if not workbook_path.exists():
        raise SystemExit(
            f"Workbook not found: {workbook_path}\n"
            "Copy clues.xlsx into source/ or pass its full path as the first argument."
        )
    if args.shard_size < 25:
        raise SystemExit("--shard-size must be at least 25.")

    overrides = load_overrides(args.overrides.resolve())
    source_hash = source_sha256(workbook_path)
    print(f"Reading {workbook_path}...")

    (
        records_by_round,
        converted_cells,
        media_review_rows,
        source_stats,
    ) = inspect_workbook(workbook_path, overrides)

    round_one_candidates, round_one_excluded_rows = reconstruct_categories(
        records_by_round[1], 1
    )
    round_two_candidates, round_two_excluded_rows = reconstruct_categories(
        records_by_round[2], 2
    )

    round_one, round_one_media_excluded, round_one_duplicates = (
        filter_and_deduplicate_categories(
            round_one_candidates,
            include_high_confidence_media=args.include_high_confidence_media,
        )
    )
    round_two, round_two_media_excluded, round_two_duplicates = (
        filter_and_deduplicate_categories(
            round_two_candidates,
            include_high_confidence_media=args.include_high_confidence_media,
        )
    )
    final_clues, final_excluded, final_duplicates = filter_and_deduplicate_final(
        records_by_round["final"],
        include_high_confidence_media=args.include_high_confidence_media,
    )

    excluded_rows = round_one_excluded_rows + round_two_excluded_rows
    excluded_categories = round_one_media_excluded + round_two_media_excluded
    duplicates = round_one_duplicates + round_two_duplicates + final_duplicates

    clean_output_directories(project_root)

    round_one_manifest = write_shards(
        project_root,
        round_one,
        "round1",
        "r1",
        "round1-categories",
        args.shard_size,
    )
    round_two_manifest = write_shards(
        project_root,
        round_two,
        "round2",
        "r2",
        "round2-categories",
        args.shard_size,
    )
    final_manifest = write_shards(
        project_root,
        final_clues,
        "final",
        "fj",
        "final-clues",
        args.shard_size,
    )

    data_version = build_data_version(
        round_one,
        round_two,
        final_clues,
        overrides,
        args.include_high_confidence_media,
    )
    total_data_bytes = (
        round_one_manifest["bytes"]
        + round_two_manifest["bytes"]
        + final_manifest["bytes"]
    )

    manifest = {
        "schemaVersion": SCHEMA_VERSION,
        "converterVersion": CONVERTER_VERSION,
        "dataVersion": data_version,
        "generatedAt": datetime.now(timezone.utc).replace(microsecond=0).isoformat(),
        "source": {
            "filename": workbook_path.name,
            "sha256": source_hash,
            **source_stats,
        },
        "filters": {
            "mediaPolicy": (
                "include-high-confidence"
                if args.include_high_confidence_media
                else "exclude-high-confidence"
            ),
            "mediaPatternNames": [
                name for name, _description, _pattern in MEDIA_PATTERNS
            ],
            "manualOverrideCount": len(overrides),
        },
        "rounds": {
            "1": {
                "kind": "category-shards",
                "expectedValues": ROUND_VALUES[1],
                "categoryCount": len(round_one),
                "clueCount": len(round_one) * 5,
                **round_one_manifest,
            },
            "2": {
                "kind": "category-shards",
                "expectedValues": ROUND_VALUES[2],
                "categoryCount": len(round_two),
                "clueCount": len(round_two) * 5,
                **round_two_manifest,
            },
            "final": {
                "kind": "clue-shards",
                "expectedValue": FINAL_VALUE,
                "clueCount": len(final_clues),
                **final_manifest,
            },
        },
        "totals": {
            "regularCategoryCount": len(round_one) + len(round_two),
            "playableClueCount": len(round_one) * 5
            + len(round_two) * 5
            + len(final_clues),
            "dataBytes": total_data_bytes,
        },
    }
    write_json(project_root / "data" / "manifest.json", manifest, pretty=True)

    summary = {
        "schemaVersion": SCHEMA_VERSION,
        "converterVersion": CONVERTER_VERSION,
        "dataVersion": data_version,
        "source": manifest["source"],
        "reconstruction": {
            "round1CompleteCategoryCandidates": len(round_one_candidates),
            "round2CompleteCategoryCandidates": len(round_two_candidates),
            "regularRowsOutsideCompleteSequences": len(excluded_rows),
        },
        "playable": {
            "round1Categories": len(round_one),
            "round1Clues": len(round_one) * 5,
            "round2Categories": len(round_two),
            "round2Clues": len(round_two) * 5,
            "finalClues": len(final_clues),
            "totalClues": manifest["totals"]["playableClueCount"],
        },
        "excluded": {
            "mediaDependentCategoryInstances": len(excluded_categories),
            "finalClues": len(final_excluded),
            "exactDuplicates": len(duplicates),
        },
        "review": {
            "mediaRowsHighConfidence": sum(
                1 for row in media_review_rows if row["confidence"] == "high"
            ),
            "mediaRowsReviewOnly": sum(
                1
                for row in media_review_rows
                if row["confidence"] == "review_only"
            ),
            "convertedOrOverriddenCells": len(converted_cells),
        },
        "shards": {
            "round1": round_one_manifest,
            "round2": round_two_manifest,
            "final": final_manifest,
        },
    }
    write_json(
        project_root / "reports" / "conversion_summary.json",
        summary,
        pretty=True,
    )

    write_csv(
        project_root / "reports" / "excluded_rows.csv",
        excluded_rows,
        [
            "source_row",
            "round",
            "value",
            "category",
            "reason_codes",
            "clue",
            "response",
        ],
    )
    write_csv(
        project_root / "reports" / "excluded_category_instances.csv",
        excluded_categories,
        [
            "round",
            "start_row",
            "end_row",
            "category",
            "reason_codes",
            "media_rows",
        ],
    )
    write_csv(
        project_root / "reports" / "excluded_final_clues.csv",
        final_excluded,
        [
            "source_row",
            "round",
            "value",
            "category",
            "reason_codes",
            "clue",
            "response",
        ],
    )
    write_csv(
        project_root / "reports" / "media_review.csv",
        media_review_rows,
        [
            "source_row",
            "round",
            "value",
            "category",
            "confidence",
            "reason_codes",
            "clue",
            "response",
        ],
    )
    write_csv(
        project_root / "reports" / "converted_cells.csv",
        converted_cells,
        [
            "source_row",
            "column",
            "data_type",
            "number_format",
            "original_value",
            "display_text",
            "override_applied",
            "override_reason",
        ],
    )
    write_csv(
        project_root / "reports" / "duplicate_records.csv",
        duplicates,
        [
            "kind",
            "id",
            "category",
            "first_source_rows",
            "duplicate_source_rows",
        ],
    )

    write_markdown_report(
        project_root / "docs" / "DATA_CONVERSION_REPORT.md",
        workbook_path,
        source_hash,
        source_stats,
        len(round_one_candidates),
        len(round_two_candidates),
        round_one,
        round_two,
        final_clues,
        excluded_rows,
        excluded_categories,
        final_excluded,
        duplicates,
        media_review_rows,
        converted_cells,
        manifest,
    )

    print("Conversion complete.")
    print(f"  Round 1 categories: {len(round_one):,}")
    print(f"  Round 2 categories: {len(round_two):,}")
    print(f"  Final clues: {len(final_clues):,}")
    print(f"  Playable clues: {manifest['totals']['playableClueCount']:,}")
    print(f"  Data size: {format_bytes(total_data_bytes)}")
    print(f"  Data version: {data_version}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except ConversionError as error:
        print(f"Conversion failed: {error}", file=sys.stderr)
        raise SystemExit(1)
